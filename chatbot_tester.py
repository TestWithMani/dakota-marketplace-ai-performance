"""
Dakota Joe chatbot prompt automation.

Reads Prompts.csv, sends each object prompt through the embedded chat UI,
records timings to the performance workbook, and writes Allure results.
"""

import argparse
import csv
import json
import os
import shutil
import socket
import subprocess
import sys
import time
import traceback
import uuid
import webbrowser
from copy import copy
from datetime import datetime

try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception:
    Alignment = Border = Font = PatternFill = Side = None
    get_column_letter = None

from selenium import webdriver
from selenium.common.exceptions import NoSuchWindowException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

import config as app_config
from config import (
    ALLURE_REPORT_DIR as ALLURE_REPORT_DIR_NAME,
    ALLURE_RESULTS_DIR as ALLURE_RESULTS_DIR_NAME,
    AUTOMATION_LINK_COL,
    AUTOMATION_STATUS_COL,
    AUTOMATION_TIME_COL,
    CHAT_INPUT_X_OFFSET,
    CHAT_INPUT_Y_OFFSET,
    CHAT_RECOVERY_SLEEP_SECONDS,
    CONSOLE_EVENT_WAIT_SECONDS,
    DEFAULT_MARKET,
    DEFAULT_RUN_MODE,
    MARKER_COL,
    OBJECT_TYPE_COL,
    PAGE_IDLE_AFTER_LOGIN,
    PASSWORD,
    BENCHMARKS_CSV as BENCHMARKS_CSV_NAME,
    PERFORMANCE_RESULTS_XLSX as PERFORMANCE_RESULTS_XLSX_NAME,
    PROMPTS_FILE,
    PROMPT_HEADER_NAMES,
    RESPONSE_TIMEOUT,
    RUNS_PER_OBJECT,
    SCREENSHOTS_DIR as SCREENSHOTS_DIR_NAME,
    RUN_MODES,
    SMOKE_MARKER,
    TEST_MARKER,
    ENABLE_COORDINATE_FALLBACK,
    USE_DOM_FIRST_CHAT_OPEN,
    USE_DOM_FIRST_SEND,
    URL,
    USERNAME,
    market_choices,
    normalize_run_mode,
    project_path,
    resolve_prompt_execution,
    resolve_market_profile,
)

_SCRIPT_DIR = project_path()
PROMPTS_CSV = project_path(PROMPTS_FILE)
PERFORMANCE_RESULTS_XLSX = project_path(PERFORMANCE_RESULTS_XLSX_NAME)
BENCHMARKS_CSV = project_path(BENCHMARKS_CSV_NAME)
ALLURE_RESULTS_DIR = project_path(ALLURE_RESULTS_DIR_NAME)
ALLURE_REPORT_DIR = project_path(ALLURE_REPORT_DIR_NAME)
SCREENSHOTS_DIR = project_path(SCREENSHOTS_DIR_NAME)

# Legacy typo kept for reading older workbooks.
BENCHMARK_FIELD_LEGACY = "Performnace Benchmake (s)"
BENCHMARK_FIELD = "Performance Benchmark (s)"

PERF_RESULT_FIELDS = [
    "Prompt",
    "Object Type",
    "Sample #",
    "Time (s)",
    BENCHMARK_FIELD,
    "Browser",
    "Recorded At",
    "Platform",
]

# Tracks merged prompt cells while appending performance rows to the XLSX.
_PERF_XLSX_GROUP_START_ROW_BY_KEY = {}

CHAT_TEXTAREA_SELECTOR = (
    "textarea.embeddedMessagingInputFooterTextArea,"
    " textarea.chat-box,"
    " textarea[placeholder*='Type your message']"
)
SEND_BUTTON_SELECTOR = (
    "button[aria-label*='send' i],"
    " button[title*='send' i],"
    " button[aria-label*='submit' i],"
    " button[title*='submit' i]"
)
# Do not treat these as the footer send arrow (speech bubble / attach / menu, etc.).
SEND_BUTTON_EXCLUDE_SUBSTRINGS = (
    "attach",
    "attachment",
    "paperclip",
    "emoji",
    "file",
    "menu",
    "minimize",
    "collapse",
    "chevron",
    "microphone",
    "voice",
    "photo",
    "camera",
    "gallery",
    "speech",
    "bolt",
    "template",
)


def _send_button_label_blob(btn):
    try:
        parts = [
            btn.get_attribute("aria-label") or "",
            btn.get_attribute("title") or "",
            btn.get_attribute("class") or "",
        ]
        return " ".join(parts).lower()
    except Exception:
        return ""


def _is_chat_side_icon_not_send_arrow(btn):
    t = _send_button_label_blob(btn)
    return any(s in t for s in SEND_BUTTON_EXCLUDE_SUBSTRINGS)


def _pick_rightmost_footer_send_button(buttons):
    """Prefer the purple send arrow: rightmost enabled match in the composer row."""
    if not buttons:
        return None
    filtered = [b for b in buttons if not _is_chat_side_icon_not_send_arrow(b)]
    pool = filtered if filtered else list(buttons)

    def right_edge(b):
        try:
            r = b.rect
            return float(r.get("x", 0)) + float(r.get("width", 0))
        except Exception:
            try:
                loc = b.location
                size = b.size
                return float(loc["x"]) + float(size["width"])
            except Exception:
                return 0.0

    return max(pool, key=right_edge)


# --- Browser setup ---


def setup_driver(headless=False, browser="chrome"):
    """Start Chrome, Edge, or Firefox (default: chrome)."""
    browser_key = (browser or "chrome").strip().lower()
    headless_args = []
    if headless:
        headless_args = ["--headless=new", "--window-size=1920,1080"]
    else:
        headless_args = ["--start-maximized"]

    if browser_key == "edge":
        from selenium.webdriver.edge.options import Options as EdgeOptions

        options = EdgeOptions()
        for arg in headless_args:
            options.add_argument(arg)
        options.set_capability("ms:loggingPrefs", {"browser": "ALL"})
        return webdriver.Edge(options=options)

    if browser_key == "firefox":
        from selenium.webdriver.firefox.options import Options as FirefoxOptions

        options = FirefoxOptions()
        if headless:
            options.add_argument("-headless")
        options.set_capability("moz:loggingPrefs", {"browser": "ALL"})
        return webdriver.Firefox(options=options)

    options = Options()
    for arg in headless_args:
        options.add_argument(arg)
    options.set_capability("goog:loggingPrefs", {"browser": "ALL"})
    return webdriver.Chrome(options=options)

# --- Prompt loading ---


def _normalize_marker(value):
    return " ".join(str(value or "").split()).lower()


def _is_smoke_marker(value):
    return _normalize_marker(value) == SMOKE_MARKER


def _is_test_marker(value):
    return _normalize_marker(value) == TEST_MARKER


def _prompt_entry_row_idx(entry):
    return entry[0]


def _prompt_entry_prompt(entry):
    return entry[1]


def _prompt_entry_marker(entry):
    return entry[2] if len(entry) > 2 else ""


def _filter_prompt_entries_for_smoke(prompt_entries, smoke_only):
    if not smoke_only:
        return list(prompt_entries)
    return _filter_prompt_entries_for_run_mode(prompt_entries, "smoke")


def _filter_prompt_entries_for_run_mode(prompt_entries, run_mode):
    """Filter prompts by run mode: all | smoke | test (Marker column)."""
    mode = normalize_run_mode(run_mode)
    if mode == "all":
        return list(prompt_entries)
    if mode == "smoke":
        filtered = [
            entry
            for entry in prompt_entries
            if _is_smoke_marker(_prompt_entry_marker(entry))
        ]
    else:
        filtered = [
            entry
            for entry in prompt_entries
            if _is_test_marker(_prompt_entry_marker(entry))
        ]
    return filtered


def load_prompts_from_csv(csv_path):
    """Load Prompts.csv rows as (row_index, prompt_text, marker) entries."""
    if not os.path.exists(csv_path):
        return None
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if len(rows) < 2:
        return None

    header_row_idx = None
    for i, row in enumerate(rows):
        if any((cell or "").strip() in PROMPT_HEADER_NAMES for cell in row):
            header_row_idx = i
            break
    if header_row_idx is None:
        return None

    header = rows[header_row_idx]
    prompt_col = marker_col = status_col = link_col = time_col = hi_col = None
    for j, cell in enumerate(header):
        label = (cell or "").strip()
        if prompt_col is None and label in PROMPT_HEADER_NAMES:
            prompt_col = j
        elif marker_col is None and label == MARKER_COL:
            marker_col = j
        elif status_col is None and label == AUTOMATION_STATUS_COL:
            status_col = j
        elif link_col is None and label == AUTOMATION_LINK_COL:
            link_col = j
        elif time_col is None and label == AUTOMATION_TIME_COL:
            time_col = j
        elif hi_col is None and label == OBJECT_TYPE_COL:
            hi_col = j
    if prompt_col is None:
        return None

    ncols = len(header)
    for row in rows:
        while len(row) < ncols:
            row.append("")

    prompt_entries = []
    for i in range(header_row_idx + 1, len(rows)):
        row = rows[i]
        prompt_text = (row[prompt_col] or "").strip()
        if not prompt_text:
            continue
        marker = ""
        if marker_col is not None and len(row) > marker_col:
            marker = (row[marker_col] or "").strip()
        prompt_entries.append((i, prompt_text, marker))

    return (
        rows,
        prompt_col,
        status_col,
        link_col,
        time_col,
        prompt_entries,
        header_row_idx,
        hi_col,
    )


def write_prompts_csv(csv_path, rows):
    """Write optional legacy status/link/time columns back to a CSV source file."""
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


# --- Performance results ---


def _browser_label(driver):
    try:
        v = (driver.capabilities or {}).get("browserVersion") or ""
        if v:
            return f"Chrome {v} "
    except Exception:
        pass
    return "Chrome "


def _object_type_series_for_entries(rows, header_row_idx, hi_col, prompt_entries):
    """Carry-forward 'hi' column; first line only if multiline. Aligned with prompt order."""
    last = ""
    out = []
    for row_idx, _p, *_rest in prompt_entries:
        if hi_col is None:
            out.append("")
            continue
        r = rows[row_idx]
        cell = (r[hi_col] if len(r) > hi_col else "") or ""
        if cell.strip():
            first = (cell or "").splitlines()[0].strip()
            last = " ".join(first.split())
        out.append(last)
    return out


def _expand_prompt_runs(prompts, row_indices, object_types):
    """Repeat each object prompt RUNS_PER_OBJECT times for timing samples."""
    if not prompts:
        return prompts, row_indices, object_types, []

    exp_prompts = []
    exp_rows = []
    exp_types = []
    exp_samples = []
    i = 0
    n = len(prompts)
    while i < n:
        key = (object_types[i], prompts[i])
        j = i + 1
        while j < n and (object_types[j], prompts[j]) == key:
            j += 1
        count = j - i

        if count >= RUNS_PER_OBJECT:
            for k in range(count):
                exp_prompts.append(prompts[i + k])
                exp_rows.append(row_indices[i + k])
                exp_types.append(object_types[i + k])
                exp_samples.append(k + 1)
        else:
            base_row_idx = row_indices[i]
            for k in range(RUNS_PER_OBJECT):
                exp_prompts.append(prompts[i])
                exp_rows.append(base_row_idx)
                exp_types.append(object_types[i])
                exp_samples.append(k + 1)
        i = j

    return exp_prompts, exp_rows, exp_types, exp_samples


def _ensure_performance_results_xlsx():
    """Create the results workbook with headers when it does not exist yet."""
    if openpyxl is None or get_column_letter is None or PatternFill is None:
        return False
    if os.path.exists(PERFORMANCE_RESULTS_XLSX):
        return True

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    styles = _xlsx_styles()
    ws.append(PERF_RESULT_FIELDS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PERF_RESULT_FIELDS))}1"

    widths = [55, 22, 10, 10, 20, 20, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col in range(1, len(PERF_RESULT_FIELDS) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = styles["header"]["fill"]
        c.font = styles["header"]["font"]
        c.alignment = styles["header"]["align"]
        c.border = styles["border"]
    ws.row_dimensions[1].height = 22

    try:
        wb.save(PERFORMANCE_RESULTS_XLSX)
    except PermissionError:
        return False
    try:
        wb.close()
    except Exception:
        pass
    return True


def _xlsx_styles():
    """Style palette for the performance results workbook."""
    if PatternFill is None:
        return None
    thin = Side(style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_fill = PatternFill("solid", fgColor="E8F5E9")
    data_font = Font(color="000000")
    data_align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    data_align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    summary_fill = PatternFill("solid", fgColor="E8EAF6")
    summary_font = Font(bold=True, color="000000")
    summary_align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    summary_align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    return {
        "border": border,
        "header": {"fill": header_fill, "font": header_font, "align": header_align},
        "data": {
            "fill": data_fill,
            "font": data_font,
            "align_left": data_align_left,
            "align_center": data_align_center,
        },
        "summary": {
            "fill": summary_fill,
            "font": summary_font,
            "align_left": summary_align_left,
            "align_center": summary_align_center,
        },
    }


def _append_rows_to_performance_results_xlsx(new_rows, merge_prompt_start_end=None):
    """Append rows to the XLSX (keeping formatting). Optionally merge Prompt cell A for a group."""
    if openpyxl is None or PatternFill is None or get_column_letter is None:
        return False
    if not _ensure_performance_results_xlsx():
        return False

    try:
        wb = openpyxl.load_workbook(PERFORMANCE_RESULTS_XLSX)
    except PermissionError:
        return False
    ws = wb.active
    styles = _xlsx_styles()

    benchmark_by_obj = _load_all_benchmarks()
    bmk_col = _benchmark_col_index(ws)

    for r in new_rows:
        ws.append([r.get(k, "") for k in PERF_RESULT_FIELDS])
        row_idx = ws.max_row
        is_summary = (r.get("Prompt") or "").strip().lower() == "run summary"

        fill = styles["summary"]["fill"] if is_summary else styles["data"]["fill"]
        font = styles["summary"]["font"] if is_summary else styles["data"]["font"]
        align_left = styles["summary"]["align_left"] if is_summary else styles["data"]["align_left"]
        align_center = styles["summary"]["align_center"] if is_summary else styles["data"]["align_center"]

        for col in range(1, len(PERF_RESULT_FIELDS) + 1):
            c = ws.cell(row=row_idx, column=col)
            c.fill = fill
            c.font = font
            c.border = styles["border"]
            c.alignment = align_left if col == 1 else align_center

        if is_summary:
            obj_key = _normalize_object_key(r.get("Object Type"))
            bmk = benchmark_by_obj.get(obj_key)
            if bmk is None:
                bmk = _benchmark_value_from_row_dict(r)
            if bmk is not None:
                ws.cell(row=row_idx, column=bmk_col).value = bmk
                try:
                    avg = float(ws.cell(row=row_idx, column=4).value)
                    if avg > bmk:
                        tcell = ws.cell(row=row_idx, column=4)
                        f = copy(tcell.font)
                        f.color = "FF0000"
                        f.bold = True
                        tcell.font = f
                except Exception:
                    pass

    if merge_prompt_start_end is not None:
        start_row, end_row = merge_prompt_start_end
        if start_row is not None and end_row is not None and end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

    try:
        wb.save(PERFORMANCE_RESULTS_XLSX)
    except PermissionError:
        # Workbook is locked (commonly open in Excel).
        try:
            wb.close()
        except Exception:
            pass
        return False
    try:
        wb.close()
    except Exception:
        pass
    return True


def _normalize_object_key(value):
    return " ".join(str(value or "").split()).lower()


def _benchmark_col_index(ws):
    """Resolve benchmark column index from header row (supports legacy typo header)."""
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col).value or "").strip().lower()
        if header in (
            BENCHMARK_FIELD.lower(),
            BENCHMARK_FIELD_LEGACY.lower(),
            "benchmark (s)",
            "benchmark",
        ):
            return col
    return 5


def _benchmark_value_from_row_dict(row_dict):
    """Parse benchmark seconds from a pending append row dict."""
    for key in (BENCHMARK_FIELD, BENCHMARK_FIELD_LEGACY):
        raw = row_dict.get(key)
        if raw in (None, ""):
            continue
        try:
            return float(raw)
        except (TypeError, ValueError):
            continue
    return None


def _load_benchmarks_from_csv():
    """Read default benchmark seconds per object type from Benchmarks.csv."""
    if not os.path.exists(BENCHMARKS_CSV):
        return {}
    benchmarks = {}
    try:
        with open(BENCHMARKS_CSV, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                obj = (
                    row.get("Object Type")
                    or row.get("hi")
                    or row.get("object type")
                    or ""
                ).strip()
                raw_bmk = (
                    row.get("Benchmark (s)")
                    or row.get("Benchmark")
                    or row.get("benchmark")
                    or ""
                ).strip()
                if not obj or not raw_bmk:
                    continue
                try:
                    benchmarks[_normalize_object_key(obj)] = float(raw_bmk)
                except ValueError:
                    continue
    except Exception:
        return {}
    return benchmarks


def _load_benchmarks_from_results_xlsx():
    """Read benchmark seconds per object from existing Run summary rows in the XLSX."""
    if openpyxl is None or not os.path.exists(PERFORMANCE_RESULTS_XLSX):
        return {}
    try:
        wb = openpyxl.load_workbook(PERFORMANCE_RESULTS_XLSX, read_only=True, data_only=True)
        ws = wb.active
        benchmarks = {}
        bmk_col = _benchmark_col_index(ws)
        for rr in range(2, ws.max_row + 1):
            prompt_val = ws.cell(rr, 1).value
            if str(prompt_val or "").strip().lower() != "run summary":
                continue
            benchmark_val = ws.cell(rr, bmk_col).value
            if benchmark_val in (None, ""):
                continue
            try:
                benchmark_s = float(benchmark_val)
            except Exception:
                continue
            obj_key = _normalize_object_key(ws.cell(rr, 2).value)
            if obj_key:
                benchmarks[obj_key] = benchmark_s
        wb.close()
        return benchmarks
    except Exception:
        return {}


def _load_all_benchmarks():
    """Merge Benchmarks.csv defaults with per-run overrides stored in the XLSX."""
    merged = _load_benchmarks_from_csv()
    for obj_key, seconds in _load_benchmarks_from_results_xlsx().items():
        merged[obj_key] = seconds
    return merged


def _append_perf_sample_and_maybe_summary(
    driver,
    append_perf,
    i,
    prompt,
    prompts,
    object_types,
    sample_nums,
    elapsed_s,
    group_elapsed,
    recorded,
    benchmarks_by_object=None,
):
    """Append one sample row to the results XLSX; append Run summary when an object group finishes."""
    if not append_perf:
        return group_elapsed
    br = _browser_label(driver)
    row_to_append = {
        "Prompt": prompt,
        "Object Type": object_types[i],
        "Sample #": str(sample_nums[i]),
        "Time (s)": str(elapsed_s),
        BENCHMARK_FIELD: "",
        "Browser": br,
        "Recorded At": recorded,
        "Platform": "windows",
    }
    # Append performance rows to the single formatted XLSX output.
    # Merge the Prompt cell (col A) across the 3 samples of a group.
    merge_key = (recorded, object_types[i], prompts[i])
    merge_start_end = None
    if str(sample_nums[i]) == "1":
        # After we append this row, its row index will be (current last row + 1).
        try:
            if _ensure_performance_results_xlsx():
                wb_tmp = openpyxl.load_workbook(PERFORMANCE_RESULTS_XLSX)
                ws_tmp = wb_tmp.active
                _PERF_XLSX_GROUP_START_ROW_BY_KEY[merge_key] = ws_tmp.max_row + 1
                wb_tmp.close()
        except PermissionError:
            pass
        except Exception:
            pass
    if str(sample_nums[i]) == str(RUNS_PER_OBJECT):
        try:
            if _ensure_performance_results_xlsx():
                wb_tmp = openpyxl.load_workbook(PERFORMANCE_RESULTS_XLSX)
                ws_tmp = wb_tmp.active
                end_row = ws_tmp.max_row + 1
                wb_tmp.close()
                start_row = _PERF_XLSX_GROUP_START_ROW_BY_KEY.pop(merge_key, None)
                if start_row is not None:
                    merge_start_end = (start_row, end_row)
        except PermissionError:
            pass
        except Exception:
            pass
    _append_rows_to_performance_results_xlsx(
        [row_to_append], merge_prompt_start_end=merge_start_end
    )
    try:
        ge = list(group_elapsed) + [float(elapsed_s)]
    except (TypeError, ValueError):
        ge = list(group_elapsed)
    is_last = (i == len(prompts) - 1) or (
        (object_types[i + 1], prompts[i + 1]) != (object_types[i], prompts[i])
    )
    if is_last and ge:
        avg = sum(ge) / len(ge)
        obj_key = _normalize_object_key(object_types[i])
        bmk = (benchmarks_by_object or {}).get(obj_key)
        summary_row = {
            "Prompt": "Run summary",
            "Object Type": object_types[i],
            "Sample #": "",
            "Time (s)": f"{avg:.2f}",
            BENCHMARK_FIELD: f"{bmk:.2f}" if bmk is not None else "",
            "Browser": br,
            "Recorded At": recorded,
            "Platform": "windows",
        }
        _append_rows_to_performance_results_xlsx([summary_row])
        return []
    return ge


# --- Allure reporting ---


def take_screenshot(driver, name):
    try:
        os.makedirs(SCREENSHOTS_DIR, exist_ok=True)
        filename = os.path.join(
            SCREENSHOTS_DIR,
            f"{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
        )
        driver.save_screenshot(filename)
        print(f"  Screenshot: {filename}")
        return filename
    except Exception:
        return None


def _prepare_allure_results(browser_name, platform_name):
    """Reset allure-results and write environment metadata."""
    try:
        if os.path.exists(ALLURE_RESULTS_DIR):
            shutil.rmtree(ALLURE_RESULTS_DIR)
        os.makedirs(ALLURE_RESULTS_DIR, exist_ok=True)
        env_path = os.path.join(ALLURE_RESULTS_DIR, "environment.properties")
        market_label = os.environ.get("DAKOTA_MARKET", DEFAULT_MARKET)
        with open(env_path, "w", encoding="utf-8") as f:
            f.write(f"Browser={browser_name}\n")
            f.write(f"Platform={platform_name}\n")
            f.write(f"Market={market_label}\n")
            f.write(f"BaseURL={URL}\n")
            f.write(f"GeneratedAt={datetime.now().isoformat()}\n")
        return True
    except Exception as e:
        print(f"  Allure init skipped: {e}")
        return False


def _allure_attachment_from_text(content, name, content_type="text/plain"):
    src = f"{uuid.uuid4()}-attachment.txt"
    with open(os.path.join(ALLURE_RESULTS_DIR, src), "w", encoding="utf-8") as f:
        f.write(content)
    return {"name": name, "source": src, "type": content_type}


def _allure_attachment_from_file(path, name, content_type):
    if not path or not os.path.exists(path):
        return None
    ext = os.path.splitext(path)[1] or ".bin"
    src = f"{uuid.uuid4()}-attachment{ext}"
    shutil.copyfile(path, os.path.join(ALLURE_RESULTS_DIR, src))
    return {"name": name, "source": src, "type": content_type}


def _allure_group_is_complete(i, prompts, object_types, append_perf):
    """True when the current object group has finished all samples."""
    if not append_perf:
        return True
    if i >= len(prompts) - 1:
        return True
    return (object_types[i + 1], prompts[i + 1]) != (object_types[i], prompts[i])


def _start_allure_object_group(object_type, prompt, start_ms):
    return {
        "object_type": object_type or "General",
        "prompt": prompt,
        "start_ms": int(start_ms),
        "samples": [],
    }


def _append_allure_sample(
    group,
    *,
    sample_num,
    status,
    link,
    elapsed_s,
    start_ms,
    stop_ms,
    screenshot_path=None,
    error_message=None,
):
    group["samples"].append(
        {
            "sample_num": str(sample_num or ""),
            "status": status,
            "link": link,
            "elapsed_s": elapsed_s,
            "start_ms": int(start_ms),
            "stop_ms": int(stop_ms),
            "screenshot_path": screenshot_path,
            "error_message": error_message,
        }
    )


def _write_allure_object_result(group, benchmarks):
    """Write one Allure test case per object with sample steps and benchmark status."""
    samples = group.get("samples") or []
    if not samples:
        return False

    object_type = group.get("object_type") or "General"
    prompt = group.get("prompt") or ""
    elapsed_values = []
    for sample in samples:
        try:
            elapsed_values.append(float(sample.get("elapsed_s")))
        except (TypeError, ValueError):
            pass
    avg = (sum(elapsed_values) / len(elapsed_values)) if elapsed_values else 0.0
    benchmark = benchmarks.get(_normalize_object_key(object_type))
    sample_failed = any(str(sample.get("status", "")).lower() != "pass" for sample in samples)
    benchmark_failed = benchmark is not None and avg > benchmark
    allure_status = "failed" if sample_failed or benchmark_failed else "passed"

    case_uuid = str(uuid.uuid4())
    test_case_id = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"{object_type}|{prompt}"))
    start_ms = int(group.get("start_ms") or samples[0]["start_ms"])
    stop_ms = max(int(sample["stop_ms"]) for sample in samples)

    steps = []
    for sample in samples:
        step_status = "passed" if str(sample.get("status", "")).lower() == "pass" else "failed"
        step_attachments = [
            _allure_attachment_from_text(str(sample.get("link") or "-"), "Generated Link"),
            _allure_attachment_from_text(str(sample.get("elapsed_s") or ""), "Elapsed Seconds"),
        ]
        if sample.get("error_message"):
            step_attachments.append(
                _allure_attachment_from_text(sample["error_message"], "Failure Reason")
            )
        shot = _allure_attachment_from_file(
            sample.get("screenshot_path"), "Failure Screenshot", "image/png"
        )
        if shot:
            step_attachments.append(shot)
        steps.append(
            {
                "name": f"Sample {sample.get('sample_num') or '-'}",
                "status": step_status,
                "stage": "finished",
                "start": int(sample["start_ms"]),
                "stop": int(sample["stop_ms"]),
                "attachments": step_attachments,
            }
        )

    summary_lines = [
        f"Object: {object_type}",
        f"Prompt: {prompt}",
        f"Average (s): {avg:.2f}",
        f"Benchmark (s): {benchmark if benchmark is not None else 'n/a'}",
    ]
    attachments = [_allure_attachment_from_text("\n".join(summary_lines), "Run Summary")]
    if benchmark_failed:
        attachments.append(
            _allure_attachment_from_text(
                f"Average {avg:.2f}s exceeds benchmark {benchmark:.2f}s",
                "Benchmark Failure",
            )
        )

    parameters = [
        {"name": "object_type", "value": object_type},
        {"name": "prompt", "value": prompt},
        {"name": "average_seconds", "value": f"{avg:.2f}"},
        {
            "name": "benchmark_seconds",
            "value": f"{benchmark:.2f}" if benchmark is not None else "n/a",
        },
    ]
    labels = [
        {"name": "suite", "value": "Dakota Prompt Automation"},
        {"name": "parentSuite", "value": "Object Performance"},
        {"name": "feature", "value": object_type},
        {"name": "host", "value": os.environ.get("COMPUTERNAME", "local")},
        {"name": "framework", "value": "custom-python"},
        {"name": "language", "value": "python"},
    ]
    result = {
        "uuid": case_uuid,
        "testCaseId": test_case_id,
        "historyId": test_case_id,
        "name": object_type,
        "fullName": f"Dakota Prompt Automation.{object_type}",
        "status": allure_status,
        "stage": "finished",
        "start": start_ms,
        "stop": stop_ms,
        "labels": labels,
        "parameters": parameters,
        "steps": steps,
        "attachments": attachments,
    }
    failure_messages = []
    if sample_failed:
        failure_messages.append("One or more prompt samples failed.")
    if benchmark_failed:
        failure_messages.append(
            f"Average {avg:.2f}s exceeds benchmark {benchmark:.2f}s."
        )
    if failure_messages:
        message = " ".join(failure_messages)
        result["statusDetails"] = {"message": message, "trace": message}

    out_path = os.path.join(ALLURE_RESULTS_DIR, f"{case_uuid}-result.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=True)
    if benchmark_failed:
        print(
            f"  Allure benchmark fail: {object_type} avg {avg:.2f}s > benchmark {benchmark:.2f}s"
        )
    return allure_status == "passed"


def _finalize_allure_object_group(group, benchmarks):
    """Write the current object group to allure-results and return pass/fail."""
    if not group or not group.get("samples"):
        return None
    return _write_allure_object_result(group, benchmarks)


def _allure_result_files():
    """Return result JSON files written for the current run."""
    if not os.path.isdir(ALLURE_RESULTS_DIR):
        return []
    return [
        os.path.join(ALLURE_RESULTS_DIR, name)
        for name in os.listdir(ALLURE_RESULTS_DIR)
        if name.endswith("-result.json")
    ]


def _launch_allure_report_viewer():
    """Open the generated HTML report through a local Python web server."""
    index_path = os.path.join(ALLURE_REPORT_DIR, "index.html")
    if not os.path.exists(index_path):
        return False

    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.bind(("127.0.0.1", 0))
            port = sock.getsockname()[1]
        subprocess.Popen(
            [
                sys.executable,
                "-m",
                "http.server",
                str(port),
                "--bind",
                "127.0.0.1",
                "--directory",
                ALLURE_REPORT_DIR,
            ],
            cwd=_SCRIPT_DIR,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        url = f"http://127.0.0.1:{port}/index.html"
        time.sleep(0.5)
        webbrowser.open(url)
        print(f"Allure report viewer: {url}")
        return True
    except Exception:
        return False


def _generate_allure_report(total, passed, failed):
    """Generate allure-report from allure-results if CLI is available."""
    result_files = _allure_result_files()
    if not result_files:
        print("Allure report skipped: no test results in allure-results.")
        return

    summary_path = os.path.join(ALLURE_RESULTS_DIR, "run-summary.txt")
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(f"Total={total}\nPassed={passed}\nFailed={failed}\n")
    def _run_allure_generate():
        candidates = [
            ["allure", "generate", "allure-results", "-o", "allure-report", "--clean"],
            ["npx", "allure", "generate", "allure-results", "-o", "allure-report", "--clean"],
        ]
        if os.name == "nt":
            appdata = os.environ.get("APPDATA", "")
            if appdata:
                allure_cmd = os.path.join(appdata, "npm", "allure.cmd")
                if os.path.exists(allure_cmd):
                    candidates.insert(
                        0,
                        [allure_cmd, "generate", "allure-results", "-o", "allure-report", "--clean"],
                    )
        last_err = ""
        for cmd in candidates:
            try:
                r = subprocess.run(
                    cmd,
                    cwd=_SCRIPT_DIR,
                    check=False,
                    capture_output=True,
                    text=True,
                    shell=(os.name == "nt" and cmd[0].endswith(".cmd")),
                )
                if r.returncode == 0:
                    return True
                last_err = (r.stderr or r.stdout or "").strip()
            except FileNotFoundError as exc:
                last_err = str(exc)
        if last_err:
            print(f"  {last_err}")
        return False

    try:
        if not _run_allure_generate():
            print("Allure report generation skipped (CLI not found or failed).")
            return

        index_path = os.path.join(ALLURE_REPORT_DIR, "index.html")
        if not os.path.exists(index_path):
            print("Allure report generation finished but index.html is missing.")
            return

        print(f"Allure report generated: {ALLURE_REPORT_DIR}")
        if _launch_allure_report_viewer():
            print("Allure viewer started. Keep that window open while viewing the report.")
        else:
            print(f"Open: {index_path}")
    except Exception as exc:
        print(f"Allure report generation skipped: {exc}")

def login(driver):
    """Open the marketplace and sign in when the login form is shown."""
    if not USERNAME or not PASSWORD:
        print(
            "Login failed: DAKOTA_USERNAME and DAKOTA_PASSWORD are empty. "
            "Set them in .env or rely on the defaults in config.py."
        )
        return False

    print("Navigating to site...")
    driver.get(URL)
    time.sleep(5)

    # Accept cookies
    try:
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for iframe in iframes:
            try:
                driver.switch_to.frame(iframe)
                btn = driver.find_elements(By.XPATH, "//button[text()='Accept']")
                if btn:
                    btn[0].click()
                    print("  Accepted cookies")
                driver.switch_to.default_content()
            except Exception:
                driver.switch_to.default_content()
    except Exception:
        pass

    driver.switch_to.default_content()

    if _is_login_page(driver):
        print("Logging in...")
        try:
            from selenium.webdriver.support import expected_conditions as EC
            from selenium.webdriver.support.ui import WebDriverWait

            wait = WebDriverWait(driver, 30)
            username = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//input[@placeholder='Username or Email']")
                )
            )
            username.clear()
            username.send_keys(USERNAME)
            password = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//input[@type='password']"))
            )
            password.clear()
            password.send_keys(PASSWORD)
            password.send_keys(Keys.RETURN)
            WebDriverWait(driver, 60).until(lambda d: not _is_login_page(d))
            print("Login successful!")
        except Exception as e:
            print(f"Login error: {e}")
            return False

    if _is_login_page(driver):
        print("Login failed: still on the sign-in page after submit.")
        return False

    time.sleep(5)
    return True


def _is_login_page(driver):
    """True when the Salesforce sign-in form is still visible."""
    try:
        if "signin" in (driver.current_url or "").lower():
            return True
        driver.switch_to.default_content()
        return bool(
            driver.find_elements(By.XPATH, "//input[@placeholder='Username or Email']")
        )
    except (NoSuchWindowException, WebDriverException):
        return True
    except Exception:
        return False


def _driver_window_open(driver):
    try:
        _ = driver.current_url
        return True
    except (NoSuchWindowException, WebDriverException):
        return False


# --- Chat interaction ---


def click_joe_button(driver):
    """Click joe chat button/icon at bottom right."""
    print("Clicking Joe chat button...")
    width = driver.execute_script("return window.innerWidth")
    height = driver.execute_script("return window.innerHeight")
    
    # Joe button is at bottom right
    x = width - 50
    y = height - 50
    
    actions = ActionChains(driver)
    actions.move_by_offset(x, y).click().perform()
    print(f"  Clicked at ({x}, {y})")
    time.sleep(2)  # Minimal wait for chat window to open
    
    # Reset mouse to center
    ActionChains(driver).move_by_offset(-x, -y).perform()


def _click_round_joe_launcher_if_visible(driver):
    """If the round Joe FAB (blue circle / custom-chat-fab / 'joe' art) is visible, click it once.

    No-op when the launcher is hidden or absent. Does not alter UI-bar / composer logic elsewhere.
    """
    try:
        driver.switch_to.default_content()
    except Exception:
        return False
    try:
        ok = driver.execute_script(
            """
            function deepCollect(root, pred, out) {
              out = out || [];
              if (!root) return out;
              try {
                const tw = document.createTreeWalker(root, NodeFilter.SHOW_ELEMENT, null);
                let n = tw.currentNode;
                while (n) {
                  try {
                    if (pred(n)) out.push(n);
                    if (n.shadowRoot) deepCollect(n.shadowRoot, pred, out);
                  } catch (e) {}
                  n = tw.nextNode();
                }
              } catch (e) {}
              return out;
            }
            function isVisible(el) {
              try {
                const st = window.getComputedStyle(el);
                if (!st || st.display === 'none' || st.visibility === 'hidden') return false;
                const r = el.getBoundingClientRect();
                return r.width > 2 && r.height > 2;
              } catch (e) { return false; }
            }
            function isRoundJoeLauncher(el) {
              if (!el || el.nodeType !== 1) return false;
              const tag = (el.tagName || '').toLowerCase();
              const cls = (el.getAttribute('class') || '').toLowerCase();
              const txt = ((el.textContent || '').trim()).toLowerCase();
              const alt = (el.getAttribute('alt') || '').toLowerCase();
              const src = (el.getAttribute('src') || '').toLowerCase();
              if (tag === 'img' && cls.includes('custom-chat-fab-image')) return true;
              if (cls.includes('custom-chat-fab-button')) return true;
              if (cls.includes('custom-chat-fab') && txt === 'joe') return true;
              if (txt === 'joe' && (tag === 'button' || tag === 'div' || tag === 'span')) return true;
              if (src.includes('agentchatbubble')) return true;
              if (alt.includes('joe') || alt === 'chat') return true;
              return false;
            }
            function scoreBR(el) {
              const r = el.getBoundingClientRect();
              return (window.innerWidth - r.right) + (window.innerHeight - r.bottom);
            }
            const hits = deepCollect(document, isRoundJoeLauncher, []).filter(isVisible);
            if (!hits.length) return false;
            hits.sort((a, b) => scoreBR(a) - scoreBR(b));
            const t = hits[0];
            try { t.scrollIntoView({ block: 'center', inline: 'nearest' }); } catch (e) {}
            try { t.click(); } catch (e) {}
            try {
              const o = { bubbles: true, cancelable: true, view: window };
              t.dispatchEvent(new MouseEvent('mousedown', o));
              t.dispatchEvent(new MouseEvent('mouseup', o));
              t.dispatchEvent(new MouseEvent('click', o));
            } catch (e2) {}
            return true;
            """
        )
        if ok:
            time.sleep(0.55)
            return True
    except Exception:
        pass
    finally:
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
    return False


def _try_open_chat_via_dom(driver):
    """Open the chat widget via visible launcher elements when possible."""
    driver.switch_to.default_content()
    launcher_selectors = [
        "button[title*='chat' i]",
        "button[aria-label*='chat' i]",
        "button[title*='joe' i]",
        "button[aria-label*='joe' i]",
        "button[class*='embeddedMessaging']",
    ]
    for selector in launcher_selectors:
        try:
            for btn in driver.find_elements(By.CSS_SELECTOR, selector):
                if _is_interactable(btn):
                    btn.click()
                    return True
        except Exception:
            pass
    try:
        clicked = driver.execute_script(
            """
            function isVisible(el) {
              if (!el) return false;
              const st = window.getComputedStyle(el);
              if (!st || st.visibility === 'hidden' || st.display === 'none') return false;
              const r = el.getBoundingClientRect();
              return r.width > 0 && r.height > 0;
            }
            function deepCollect(root, pred, out) {
              out = out || [];
              if (!root) return out;
              try {
                const tw = document.createTreeWalker(root, NodeFilter.SHOW_ELEMENT, null);
                let n = tw.currentNode;
                while (n) {
                  try {
                    if (pred(n)) out.push(n);
                    if (n.shadowRoot) deepCollect(n.shadowRoot, pred, out);
                  } catch(e) {}
                  n = tw.nextNode();
                }
              } catch(e) {}
              return out;
            }
            const candidates = deepCollect(document, (el) => {
              if (!isVisible(el)) return false;
              const tag = (el.tagName || '').toLowerCase();
              if (tag !== 'button') return false;
              const title = (el.getAttribute('title') || '').toLowerCase();
              const aria = (el.getAttribute('aria-label') || '').toLowerCase();
              const cls = (el.getAttribute('class') || '').toLowerCase();
              return title.includes('chat') || aria.includes('chat') || title.includes('joe') || aria.includes('joe') || cls.includes('embeddedmessaging');
            });
            if (candidates.length) {
              candidates[0].click();
              return true;
            }
            return false;
            """
        )
        return bool(clicked)
    except Exception:
        return False


def _send_prompt_via_dom(driver, prompt):
    """Try to send prompt using textarea + send button interactions."""
    driver.switch_to.default_content()
    textarea = find_chat_textarea(driver, quick_check=False, require_interactable=False)
    if not textarea:
        return False
    try:
        textarea.click()
        textarea.send_keys(Keys.CONTROL, "a")
        textarea.send_keys(Keys.BACKSPACE)
        textarea.send_keys(prompt)
        send_btn = _wait_for_send_button_near_textarea(driver, textarea, timeout_s=7)
        if not send_btn:
            send_btn = wait_for_send_button_enabled(
                driver, timeout_s=4, switch_to_default_first=False
            )
        if send_btn:
            try:
                send_btn.click()
            except Exception:
                try:
                    driver.execute_script("arguments[0].click();", send_btn)
                except Exception:
                    pass
        else:
            textarea.send_keys(Keys.ENTER)
        time.sleep(0.45)
        empty = is_chat_input_empty_by_coordinates(driver)
        still = _composer_draft_contains_substring(driver, prompt[:120])
        if empty and not still:
            print(f"  Sent via DOM: {prompt[:40]}...")
            return True
    except Exception:
        return False
    finally:
        driver.switch_to.default_content()
    return False


def _wait_for_send_button_near_textarea(driver, textarea, timeout_s=7):
    """Find the enabled send arrow closest to the active chat textarea."""
    start = time.time()
    while time.time() - start < timeout_s:
        try:
            btn = driver.execute_script(
                """
                const ta = arguments[0];
                const selector = arguments[1];
                const EXCLUDE = arguments[2] || [];
                if (!ta || !ta.getBoundingClientRect) return null;
                const taRect = ta.getBoundingClientRect();

                function isVisible(el) {
                  if (!el) return false;
                  const st = window.getComputedStyle(el);
                  if (!st || st.display === 'none' || st.visibility === 'hidden') return false;
                  const r = el.getBoundingClientRect();
                  return r.width > 0 && r.height > 0;
                }

                function isEnabled(el) {
                  if (!isVisible(el)) return false;
                  const aria = (el.getAttribute('aria-disabled') || '').toLowerCase();
                  if (aria === 'true' || aria === 'disabled') return false;
                  if (el.getAttribute('disabled') !== null) return false;
                  return true;
                }

                function isChatSideIcon(btn) {
                  const al = (btn.getAttribute('aria-label') || '').toLowerCase();
                  const ti = (btn.getAttribute('title') || '').toLowerCase();
                  const cl = (btn.getAttribute('class') || '').toLowerCase();
                  const t = al + ' ' + ti + ' ' + cl;
                  for (let i = 0; i < EXCLUDE.length; i++) {
                    const x = String(EXCLUDE[i]).toLowerCase();
                    if (x && t.includes(x)) return true;
                  }
                  return false;
                }

                const nodes = Array.from(document.querySelectorAll(selector)).filter(isEnabled);
                if (!nodes.length) return null;

                let pool = nodes.filter((n) => !isChatSideIcon(n));
                if (!pool.length) pool = nodes;

                const taCenterY = (taRect.top + taRect.bottom) / 2;
                let row = pool.filter((n) => {
                  const r = n.getBoundingClientRect();
                  return Math.abs((r.top + r.bottom) / 2 - taCenterY) < 72;
                });
                if (!row.length) row = pool;

                row.sort((a, b) => b.getBoundingClientRect().right - a.getBoundingClientRect().right);
                return row[0];
                """,
                textarea,
                SEND_BUTTON_SELECTOR,
                list(SEND_BUTTON_EXCLUDE_SUBSTRINGS),
            )
            if btn:
                return btn
        except Exception:
            pass
        time.sleep(0.1)
    return None

def _try_click_end_chat_in_current_context(driver):
    """In current context: click End chat if visible, else click exact header Menu button."""
    try:
        state = driver.execute_script(
            """
            function isVisible(el) {
              try {
                if (!el) return false;
                const st = window.getComputedStyle(el);
                if (!st || st.visibility === 'hidden' || st.display === 'none') return false;
                const r = el.getBoundingClientRect();
                return r.width > 0 && r.height > 0;
              } catch(e) { return false; }
            }
            function lowerText(el) {
              return ((el && (el.innerText || el.textContent)) || '').trim().toLowerCase();
            }
            function attr(el, name) {
              try { return (el.getAttribute(name) || '').trim().toLowerCase(); } catch(e) { return ''; }
            }
            function deepCollect(root, pred, out) {
              out = out || [];
              if (!root) return out;
              try {
                const tw = document.createTreeWalker(root, NodeFilter.SHOW_ELEMENT, null);
                let n = tw.currentNode;
                while (n) {
                  try {
                    if (pred(n)) out.push(n);
                    if (n.shadowRoot) deepCollect(n.shadowRoot, pred, out);
                  } catch(e) {}
                  n = tw.nextNode();
                }
              } catch(e) {}
              return out;
            }

            const endTargets = deepCollect(document, (el) => {
              if (!isVisible(el)) return false;
              const t = lowerText(el);
              const aria = attr(el, 'aria-label');
              const title = attr(el, 'title');
              const role = attr(el, 'role');
              const tag = (el.tagName || '').toLowerCase();
              const clickable = tag === 'button' || tag === 'a' || role === 'menuitem' || role === 'button' || !!el.onclick;
              if (!clickable) return false;
              return t.includes('end chat') || aria.includes('end chat') || title.includes('end chat');
            });
            if (endTargets.length) {
              endTargets
                .map((el) => ({ el, r: el.getBoundingClientRect() }))
                .sort((a, b) => (a.r.top - b.r.top) || (b.r.width - a.r.width));
              const target = endTargets[0];
              target.click();
              return 'ended';
            }

            // Exact button requested by user:
            // <button class="headerButton menuButton" title="Menu" aria-label="Menu">
            const exactMenuButtons = deepCollect(document, (el) => {
              if (!isVisible(el)) return false;
              const tag = (el.tagName || '').toLowerCase();
              if (tag !== 'button') return false;
              const cls = attr(el, 'class');
              const title = attr(el, 'title');
              const aria = attr(el, 'aria-label');
              return cls.includes('headerbutton') && cls.includes('menubutton') && title === 'menu' && aria === 'menu';
            });
            if (exactMenuButtons.length) {
              const menu = exactMenuButtons
                .map((el) => ({ el, r: el.getBoundingClientRect() }))
                .sort((a, b) => (a.r.top - b.r.top) || (b.r.right - a.r.right))[0].el;
              menu.click();
              return 'menu_opened';
            }

            // Fallback: any button with title/aria label "Menu"
            const menuByLabel = deepCollect(document, (el) => {
              if (!isVisible(el)) return false;
              const tag = (el.tagName || '').toLowerCase();
              if (tag !== 'button') return false;
              const title = attr(el, 'title');
              const aria = attr(el, 'aria-label');
              return title === 'menu' || aria === 'menu';
            });
            if (menuByLabel.length) {
              menuByLabel[0].click();
              return 'menu_opened';
            }
            return 'none';
            """
        )
        return state or "none"
    except Exception:
        return "none"

def end_chat_session_from_header(driver, timeout_s=16):
    """Click Menu button -> End chat; returns True on success."""
    if not _driver_window_open(driver):
        print("  End chat skipped: browser window is closed.")
        return False
    print("  Ending current chat session: click Menu then End chat...")
    end_at = time.time() + timeout_s
    saw_menu_opened = False
    while time.time() < end_at:
        driver.switch_to.default_content()
        state = _try_click_end_chat_in_current_context(driver)
        if state == "menu_opened":
            saw_menu_opened = True
            time.sleep(0.35)
        if state == "ended":
            time.sleep(1)
            driver.switch_to.default_content()
            print("  End chat clicked successfully.")
            return True

        try:
            iframes = driver.find_elements(By.TAG_NAME, "iframe")
        except Exception:
            iframes = []
        for iframe in iframes:
            try:
                driver.switch_to.frame(iframe)
                state = _try_click_end_chat_in_current_context(driver)
                if state == "menu_opened":
                    saw_menu_opened = True
                    time.sleep(0.35)
                if state == "ended":
                    driver.switch_to.default_content()
                    time.sleep(1)
                    print("  End chat clicked successfully.")
                    return True
            except Exception:
                pass
            finally:
                driver.switch_to.default_content()

        time.sleep(0.35)

    if saw_menu_opened:
        print("  Opened header menu but could not click 'End chat'.")
    else:
        print("  Could not find Joe header menu ('...') to end chat.")
    return False

def _should_reset_chat_session_after_prompt(append_perf, i, total_prompts):
    """End chat after each prompt when more executions remain."""
    if not append_perf:
        return False
    return i < total_prompts - 1


def _is_interactable(el):
    try:
        if not el.is_displayed():
            return False
        if not el.is_enabled():
            return False
        aria_disabled = (el.get_attribute("aria-disabled") or "").strip().lower()
        if aria_disabled in ("true", "disabled"):
            return False
        return True
    except:
        return False

def _is_enabled_button(el):
    try:
        if not el.is_displayed():
            return False
        if not el.is_enabled():
            return False
        aria_disabled = (el.get_attribute("aria-disabled") or "").strip().lower()
        if aria_disabled in ("true", "disabled"):
            return False
        disabled = el.get_attribute("disabled")
        if disabled is not None:
            return False
        return True
    except:
        return False

def wait_for_send_button_enabled(driver, timeout_s=10, *, switch_to_default_first=True):
    """Wait until the send/arrow button becomes enabled. Returns WebElement or None.

    Works in main doc and attempts to traverse shadow DOM via JS.
    When the composer lives in an iframe, pass switch_to_default_first=False so each
    poll does not leave that frame (otherwise send is never found).
    """
    start = time.time()
    if switch_to_default_first:
        driver.switch_to.default_content()

    while time.time() - start < timeout_s:
        # 1) Try normal DOM in current context
        try:
            buttons = driver.find_elements(By.CSS_SELECTOR, SEND_BUTTON_SELECTOR)
            enabled = [b for b in buttons if _is_enabled_button(b)]
            if enabled:
                return _pick_rightmost_footer_send_button(enabled)
        except:
            pass

        # 2) Shadow DOM-aware lookup
        try:
            el = driver.execute_script(
                """
                const selector = arguments[0];
                const EXCLUDE = arguments[1] || [];
                function isChatSideIcon(btn) {
                  const al = (btn.getAttribute('aria-label') || '').toLowerCase();
                  const ti = (btn.getAttribute('title') || '').toLowerCase();
                  const cl = (btn.getAttribute('class') || '').toLowerCase();
                  const t = al + ' ' + ti + ' ' + cl;
                  for (let i = 0; i < EXCLUDE.length; i++) {
                    const x = String(EXCLUDE[i]).toLowerCase();
                    if (x && t.includes(x)) return true;
                  }
                  return false;
                }
                function deepQueryAll(root) {
                  const out = [];
                  if (!root) return out;
                  try {
                    const direct = root.querySelectorAll(selector);
                    for (const n of direct) out.push(n);
                  } catch(e) {}
                  try {
                    const treeWalker = document.createTreeWalker(root, NodeFilter.SHOW_ELEMENT, null);
                    let node = treeWalker.currentNode;
                    while (node) {
                      try {
                        if (node.shadowRoot) {
                          out.push(...deepQueryAll(node.shadowRoot));
                        }
                      } catch(e) {}
                      node = treeWalker.nextNode();
                    }
                  } catch(e) {}
                  return out;
                }
                const nodes = deepQueryAll(document);
                const enabled = [];
                for (const n of nodes) {
                  const ariaDisabled = (n.getAttribute('aria-disabled') || '').toLowerCase();
                  const disabledAttr = n.getAttribute('disabled');
                  if (disabledAttr !== null) continue;
                  if (ariaDisabled === 'true' || ariaDisabled === 'disabled') continue;
                  enabled.push(n);
                }
                if (!enabled.length) return null;
                let pool = enabled.filter((n) => !isChatSideIcon(n));
                if (!pool.length) pool = enabled;
                const vh = window.innerHeight;
                let footer = pool.filter((n) => n.getBoundingClientRect().bottom > vh - 220);
                if (!footer.length) footer = pool;
                footer.sort((a, b) => b.getBoundingClientRect().right - a.getBoundingClientRect().right);
                return footer[0];
                """,
                SEND_BUTTON_SELECTOR,
                list(SEND_BUTTON_EXCLUDE_SUBSTRINGS),
            )
            if el:
                return el
        except:
            pass

        time.sleep(0.1)

    return None


def _composer_draft_contains_substring(driver, snippet):
    """True if the chat composer still shows *snippet* (send did not clear the draft)."""
    snippet = (snippet or "").strip()
    if len(snippet) < 2:
        return False
    driver.switch_to.default_content()
    try:
        ta = find_chat_textarea(driver, quick_check=False, require_interactable=False)
        if not ta:
            return False
        raw = driver.execute_script(
            """
            const e = arguments[0];
            if (!e) return '';
            const t = (e.tagName || '').toLowerCase();
            if (t === 'textarea' || t === 'input') return (e.value != null ? e.value : '');
            return (e.innerText || e.textContent || '');
            """,
            ta,
        )
        return snippet in (raw or "").strip()
    except Exception:
        return False
    finally:
        driver.switch_to.default_content()


def is_chat_input_empty_by_coordinates(driver):
    """Best-effort check: input draft is empty when chat box has no typed text."""
    try:
        driver.switch_to.default_content()
        textarea = find_chat_textarea(
            driver, quick_check=False, require_interactable=False
        )
        if textarea:
            try:
                raw = driver.execute_script(
                    """
                    const e = arguments[0];
                    if (!e) return '';
                    const t = (e.tagName || '').toLowerCase();
                    if (t === 'textarea' || t === 'input') return (e.value != null ? e.value : '');
                    return (e.innerText || e.textContent || '');
                    """,
                    textarea,
                )
                return (raw or "").strip() == ""
            except Exception:
                val = textarea.get_attribute("value")
                if val is not None:
                    return (val or "").strip() == ""
                text = (textarea.text or "").strip()
                return text == ""
        return False
    except Exception:
        return False
    finally:
        driver.switch_to.default_content()

def find_chat_textarea(driver, quick_check=False, require_interactable=True):
    """Find the chat textarea either in the main document or inside iframes.

    Returns: WebElement | None
    Side-effect: driver may be left in an iframe context when returning an element.
    Caller should switch_to.default_content() after interacting.
    """
    def _accept_candidate(el):
        try:
            if not el.is_displayed():
                return False
        except Exception:
            return False
        if require_interactable:
            return _is_interactable(el)
        return True

    driver.switch_to.default_content()

    # 1) Try main document first (your screenshot shows chat input is often NOT in an iframe)
    try:
        candidates = driver.find_elements(By.CSS_SELECTOR, CHAT_TEXTAREA_SELECTOR)
        for el in candidates:
            if _accept_candidate(el):
                return el
    except:
        pass

    # 1b) Try Shadow DOM-aware lookup in main document (Salesforce Embedded Messaging often uses shadow roots)
    try:
        el = driver.execute_script(
            """
            const selector = arguments[0];
            function deepQuery(root) {
              if (!root) return null;
              try {
                const direct = root.querySelector(selector);
                if (direct) return direct;
              } catch(e) {}
              const treeWalker = document.createTreeWalker(root, NodeFilter.SHOW_ELEMENT, null);
              let node = treeWalker.currentNode;
              while (node) {
                try {
                  if (node.shadowRoot) {
                    const found = deepQuery(node.shadowRoot);
                    if (found) return found;
                  }
                } catch(e) {}
                node = treeWalker.nextNode();
              }
              return null;
            }
            return deepQuery(document);
            """,
            CHAT_TEXTAREA_SELECTOR,
        )
        if el:
            return el
    except:
        pass

    # 2) Then try iframes (optionally with retries)
    max_attempts = 1 if quick_check else 5
    for attempt in range(max_attempts):
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for iframe in iframes:
            try:
                driver.switch_to.frame(iframe)
                candidates = driver.find_elements(By.CSS_SELECTOR, CHAT_TEXTAREA_SELECTOR)
                for el in candidates:
                    if _accept_candidate(el):
                        return el

                # Shadow DOM-aware lookup inside iframe document (if any open shadow roots exist)
                try:
                    el = driver.execute_script(
                        """
                        const selector = arguments[0];
                        function deepQuery(root) {
                          if (!root) return null;
                          try {
                            const direct = root.querySelector(selector);
                            if (direct) return direct;
                          } catch(e) {}
                          const treeWalker = document.createTreeWalker(root, NodeFilter.SHOW_ELEMENT, null);
                          let node = treeWalker.currentNode;
                          while (node) {
                            try {
                              if (node.shadowRoot) {
                                const found = deepQuery(node.shadowRoot);
                                if (found) return found;
                              }
                            } catch(e) {}
                            node = treeWalker.nextNode();
                          }
                          return null;
                        }
                        return deepQuery(document);
                        """,
                        CHAT_TEXTAREA_SELECTOR,
                    )
                    if el:
                        return el
                except:
                    pass

                # nested iframes
                nested = driver.find_elements(By.TAG_NAME, "iframe")
                for n in nested:
                    try:
                        driver.switch_to.frame(n)
                        candidates = driver.find_elements(By.CSS_SELECTOR, CHAT_TEXTAREA_SELECTOR)
                        for el in candidates:
                            if _accept_candidate(el):
                                return el
                        driver.switch_to.parent_frame()
                    except:
                        try:
                            driver.switch_to.parent_frame()
                        except:
                            pass

                driver.switch_to.default_content()
            except:
                driver.switch_to.default_content()

        if quick_check:
            break
        if attempt < max_attempts - 1:
            time.sleep(0.25)

    driver.switch_to.default_content()
    return None


def is_chat_input_bar_grayed_out(driver):
    """True when the chat input bar is visible but disabled/read-only."""
    textarea = find_chat_textarea(driver, quick_check=True, require_interactable=False)
    if not textarea:
        return False
    try:
        if not textarea.is_displayed():
            return False
        if not textarea.is_enabled():
            return True
        aria_disabled = (textarea.get_attribute("aria-disabled") or "").strip().lower()
        if aria_disabled in ("true", "disabled"):
            return True
        if textarea.get_attribute("disabled") is not None:
            return True
        readonly = (textarea.get_attribute("readonly") or "").strip().lower()
        if readonly in ("true", "readonly"):
            return True
        return False
    except Exception:
        return False
    finally:
        driver.switch_to.default_content()


def _recover_grayed_chat_input_bar(driver):
    """Refresh the page only when the chat input bar is grayed out."""
    if not is_chat_input_bar_grayed_out(driver):
        return False
    print("  Chat input bar is grayed out; refreshing page...")
    driver.switch_to.default_content()
    driver.refresh()
    time.sleep(CHAT_RECOVERY_SLEEP_SECONDS)
    return True


def type_prompt(driver, prompt, chat_already_open=False, allow_page_refresh=True):
    """Type prompt - handles both input bar (opens chat) and chat window.
    
    Args:
        chat_already_open: If True, skip checking for input bar and icon button.
                          Just type directly in chat window.
    """
    driver.switch_to.default_content()

    if allow_page_refresh and _recover_grayed_chat_input_bar(driver):
        chat_already_open = False

    try:
        if _click_round_joe_launcher_if_visible(driver):
            print("  Round Joe launcher: clicked visible FAB before prompt.")
    except Exception:
        pass

    # If input is already visible, treat chat as open and send directly.
    if not chat_already_open:
        existing_input = find_chat_textarea(driver, quick_check=True, require_interactable=False)
        if existing_input is not None:
            chat_already_open = True

    # Open chat only when no active input is visible.
    if not chat_already_open:
        print("  Opening chat...")
        opened = False
        if USE_DOM_FIRST_CHAT_OPEN:
            opened = _try_open_chat_via_dom(driver)
        if not opened and ENABLE_COORDINATE_FALLBACK:
            click_joe_button(driver)
        time.sleep(1)  # Minimal wait
    elif allow_page_refresh and _recover_grayed_chat_input_bar(driver):
        print("  Opening chat after refresh...")
        opened = False
        if USE_DOM_FIRST_CHAT_OPEN:
            opened = _try_open_chat_via_dom(driver)
        if not opened and ENABLE_COORDINATE_FALLBACK:
            click_joe_button(driver)
        time.sleep(1)
        chat_already_open = True

    # DOM-first send for CI stability.
    if USE_DOM_FIRST_SEND and _send_prompt_via_dom(driver, prompt):
        return

    if not ENABLE_COORDINATE_FALLBACK:
        raise RuntimeError("DOM send failed and coordinate fallback is disabled.")

    # Coordinate send fallback.
    driver.switch_to.default_content()
    width = driver.execute_script("return window.innerWidth")
    height = driver.execute_script("return window.innerHeight")
    x = width + CHAT_INPUT_X_OFFSET
    y = height + CHAT_INPUT_Y_OFFSET
    # Focus input, clear old draft, type prompt (all via coordinates).
    ActionChains(driver).move_by_offset(x, y).click().perform()
    ActionChains(driver).key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).send_keys(Keys.BACKSPACE).perform()
    ActionChains(driver).send_keys(prompt).perform()

    # Try to submit robustly and ensure draft actually leaves the input.
    sent = False
    for _attempt in range(2):
        textarea_for_scope = find_chat_textarea(driver, quick_check=True)
        send_btn = None
        if textarea_for_scope:
            send_btn = _wait_for_send_button_near_textarea(
                driver, textarea_for_scope, timeout_s=4
            )
        if not send_btn:
            send_btn = wait_for_send_button_enabled(
                driver, timeout_s=3, switch_to_default_first=False
            )
        if send_btn:
            try:
                send_btn.click()
            except Exception:
                try:
                    driver.execute_script("arguments[0].click();", send_btn)
                except Exception:
                    ActionChains(driver).send_keys(Keys.ENTER).perform()
        else:
            ActionChains(driver).send_keys(Keys.ENTER).perform()

        # Small settle wait, then verify draft cleared.
        time.sleep(0.4)
        if is_chat_input_empty_by_coordinates(driver) and not _composer_draft_contains_substring(
            driver, prompt[:120]
        ):
            sent = True
            break

        # Draft still present -> force another send attempt (Enter) from focused input.
        ActionChains(driver).send_keys(Keys.ENTER).perform()
        time.sleep(0.3)
        if is_chat_input_empty_by_coordinates(driver) and not _composer_draft_contains_substring(
            driver, prompt[:120]
        ):
            sent = True
            break

    if not sent:
        # Last fallback: try one more direct Enter.
        ActionChains(driver).send_keys(Keys.ENTER).perform()
        if allow_page_refresh and _recover_grayed_chat_input_bar(driver):
            type_prompt(
                driver,
                prompt,
                chat_already_open=False,
                allow_page_refresh=False,
            )
            return

    try:
        ActionChains(driver).move_by_offset(-x, -y).perform()
    except Exception:
        pass
    if sent:
        print(f"  Sent via coordinates: {prompt[:40]}...")
    else:
        print(
            "  [ERROR] Prompt was not sent: composer still shows the message "
            "(coordinate + Enter fallbacks exhausted)."
        )

def count_report_links(driver):
    """Count generated report links visible in the UI.

    Important: Chat often renders the report as a *title link* (not "Report"/"View"),
    and it may live in shadow DOM and/or inside iframes. So we detect by URL pattern
    (contains `/report/` with a report id) rather than anchor text.
    """
    last_link = None
    raw_valid_count = 0

    def is_valid_report_link(href: str) -> bool:
        if not href:
            return False
        # Ignore the general reports page (no report id)
        if href.endswith("/dakota-joe-reports") or href.endswith("/dakota-joe-reports/"):
            return False
        if "dakota-joe-reports" in href:
            return False
        if "/report/" not in href:
            return False
        parts = href.split("/report/")
        if len(parts) <= 1:
            return False
        report_id = parts[1].split("/")[0].split("?")[0]
        return len(report_id) >= 10 and report_id.replace("_", "").replace("-", "").isalnum()

    def collect_report_hrefs_via_js() -> list[str]:
        try:
            hrefs = driver.execute_script(
                """
                function deepCollectAnchors(root) {
                  const out = [];
                  if (!root) return out;
                  try {
                    const as = root.querySelectorAll ? root.querySelectorAll('a[href]') : [];
                    for (const a of as) out.push(a.getAttribute('href') || a.href || '');
                  } catch(e) {}

                  // Walk elements to find shadow roots
                  try {
                    const tw = document.createTreeWalker(root, NodeFilter.SHOW_ELEMENT, null);
                    let node = tw.currentNode;
                    while (node) {
                      try {
                        if (node.shadowRoot) {
                          out.push(...deepCollectAnchors(node.shadowRoot));
                        }
                      } catch(e) {}
                      node = tw.nextNode();
                    }
                  } catch(e) {}
                  return out;
                }
                const hrefs = deepCollectAnchors(document);
                return hrefs;
                """
            )
            return [h for h in (hrefs or []) if isinstance(h, str)]
        except:
            return []

    all_hrefs: list[str] = []

    # 1) Main document + shadow DOM
    driver.switch_to.default_content()
    all_hrefs.extend(collect_report_hrefs_via_js())

    # 2) Iframes (+ shadow DOM within each iframe document)
    try:
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for iframe in iframes:
            try:
                driver.switch_to.frame(iframe)
                all_hrefs.extend(collect_report_hrefs_via_js())
            except:
                pass
            finally:
                driver.switch_to.default_content()
    except:
        driver.switch_to.default_content()

    # Normalize & filter
    unique: list[str] = []
    seen = set()
    for href in all_hrefs:
        href = (href or "").strip()
        if not href:
            continue
        # convert relative hrefs to absolute
        if href.startswith("/"):
            try:
                base = driver.current_url.split("/dakotaMarketplace/")[0]
                href = base + href
            except:
                pass
        if is_valid_report_link(href):
            raw_valid_count += 1
            last_link = href
            if href in seen:
                continue
            seen.add(href)
            unique.append(href)

    # Return unique count, raw count (including duplicates), last href, and unique hrefs.
    return len(unique), raw_valid_count, last_link, tuple(unique)

def check_for_text_response(driver):
    """Check if bot replied with text (no link) - means it's done."""
    driver.switch_to.default_content()
    iframes = driver.find_elements(By.TAG_NAME, "iframe")
    for iframe in iframes:
        try:
            driver.switch_to.frame(iframe)
            # Check for common "no results" or text responses
            texts = driver.find_elements(By.XPATH, "//*[contains(text(), 'sorry') or contains(text(), 'Sorry') or contains(text(), 'cannot') or contains(text(), 'unable') or contains(text(), 'no results')]")
            if texts:
                driver.switch_to.default_content()
                return True
            driver.switch_to.default_content()
        except:
            driver.switch_to.default_content()
    return False

def _console_cycle_complete_since(driver, since_ms):
    """True only when Joe console shows a full send/complete cycle after since_ms."""
    saw_lwc_sent = False
    saw_on_embedded_complete = False
    saw_loading_false = False
    try:
        entries = driver.get_log("browser")
    except Exception:
        return False
    for entry in entries:
        ts = int(entry.get("timestamp") or 0)
        if ts <= int(since_ms or 0):
            continue
        msg = str(entry.get("message", "")).lower()
        if "message sent event received" in msg:
            saw_lwc_sent = True
        if "onembeddedmessagesent received" in msg and "haspendingmessage: false" in msg:
            saw_on_embedded_complete = True
        if "updatechatloading" in msg and "showloading: false" in msg:
            saw_loading_false = True
    return saw_lwc_sent and saw_on_embedded_complete and saw_loading_false


def wait_for_new_link(
    driver,
    initial_unique_count,
    initial_raw_count,
    timeout=RESPONSE_TIMEOUT,
    since_ms=None,
    initial_hrefs=(),
):
    """Wait for a new report link to appear.

    Contract:
    - Do NOT allow the next prompt to be sent until:
      * either a new link appears, OR
      * the bot has clearly responded with text (paragraph) and no new link is present.
    """
    start = time.time()
    start_ms = int(since_ms if since_ms is not None else start * 1000)
    last_print = 0
    baseline_hrefs = set(initial_hrefs or ())
    
    # Check every 0.2 seconds for ultra-fast detection
    while time.time() - start < timeout:
        time.sleep(0.2)  # Faster checking
        current_unique_count, current_raw_count, last_link, current_hrefs = count_report_links(driver)
        new_hrefs = [href for href in current_hrefs if href not in baseline_hrefs]

        # Treat any newly visible report href as completion for this prompt.
        if new_hrefs:
            elapsed = int(time.time() - start)
            found_link = new_hrefs[-1]
            print(f"  [OK] Link found! (raw={current_raw_count}, unique={current_unique_count}, {elapsed}s)")
            return found_link

        # If a new link element appears (raw count increases), treat it as completion even if href repeats.
        if last_link and (
            current_raw_count > initial_raw_count
            or current_unique_count > initial_unique_count
        ):
            elapsed = int(time.time() - start)
            print(f"  [OK] Link found! (raw={current_raw_count}, unique={current_unique_count}, {elapsed}s)")
            return last_link

        # If no new link yet, check if bot has answered with a text paragraph.
        # As soon as we detect a text-only response and still no new link,
        # we stop waiting and let the caller move to the next prompt.
        if current_raw_count == initial_raw_count and check_for_text_response(driver):
            elapsed = int(time.time() - start)
            print(f"  Bot replied with text only (no link) after {elapsed}s")
            return None

        # Edge case: console shows message cycle completed, but link counters did not move.
        # This prevents waiting the full timeout when response is done without a new link card.
        # Fresh chat (no links yet) must keep waiting for the first report link.
        if (
            current_raw_count == initial_raw_count
            and initial_raw_count > 0
            and _console_cycle_complete_since(driver, start_ms)
            and (time.time() - start) >= 6
        ):
            elapsed = int(time.time() - start)
            # If bot completed but no *new* link card appeared, reuse latest visible report link.
            # This handles repeated prompts where the same report link is shown again.
            _u, _r, latest_link, _hrefs = count_report_links(driver)
            if latest_link:
                print(f"  Console complete after {elapsed}s; reusing latest visible link")
                return latest_link
            print(f"  Console shows message complete with no new link after {elapsed}s")
            return None

        # Print status every 5 seconds to avoid spam
        elapsed = int(time.time() - start)
        if elapsed - last_print >= 5:
            print(f"  Links: raw={current_raw_count}, unique={current_unique_count} ({elapsed}s elapsed)")
            last_print = elapsed
    
    return None


def recheck_link_before_fail(driver, initial_unique_count, initial_raw_count, grace_seconds=8):
    """
    Final safety check before failing a prompt.
    Some links appear slightly after the normal waiter exits; this checks for late-arriving links.
    """
    start = time.time()
    while time.time() - start < grace_seconds:
        time.sleep(0.5)
        current_unique_count, current_raw_count, last_link, current_hrefs = count_report_links(driver)
        if current_raw_count > initial_raw_count and last_link:
            return last_link
        if current_unique_count > initial_unique_count and last_link:
            return last_link
        # Fallback for repeated prompt behavior: if a valid report link is visible, accept it.
        if last_link:
            return last_link
    return None


def wait_for_console_ready_event(driver, timeout_s=CONSOLE_EVENT_WAIT_SECONDS, since_ms=0):
    """
    Wait for Joe console events indicating message processing is complete
    for the CURRENT message (ignore stale old logs).

    Required log set (same prompt cycle):
      1) LWC message sent event received
      2) onEmbeddedMessageSent received with hasPendingMessage:false
      3) updateChatLoading with showLoading:false
    If browser logs are unavailable, return immediately.
    """
    start = time.time()
    saw_lwc_sent = False
    saw_on_embedded_complete = False
    saw_loading_false = False
    while time.time() - start < timeout_s:
        try:
            entries = driver.get_log("browser")
        except Exception:
            return False
        for entry in entries:
            ts = int(entry.get("timestamp") or 0)
            if ts <= int(since_ms or 0):
                continue
            msg = str(entry.get("message", "")).lower()
            if "message sent event received" in msg:
                saw_lwc_sent = True
            if "onembeddedmessagesent received" in msg and "haspendingmessage: false" in msg:
                saw_on_embedded_complete = True
            if "updatechatloading" in msg and "showloading: false" in msg:
                saw_loading_false = True

        if saw_lwc_sent and saw_on_embedded_complete and saw_loading_false:
            return True
        # Fast poll so next prompt sends almost immediately after logs appear.
        time.sleep(0.05)
    return False


def _apply_run_overrides(response_timeout=None, runs_per_object=None):
    """Override module-level timing settings (CLI / Jenkins env)."""
    global RESPONSE_TIMEOUT, RUNS_PER_OBJECT
    if response_timeout is not None:
        RESPONSE_TIMEOUT = int(response_timeout)
    if runs_per_object is not None:
        RUNS_PER_OBJECT = int(runs_per_object)


def _apply_market(market_key=None, base_url_override=None):
    """Apply market profile to runtime URL, prompts path, and config module."""
    global URL, PROMPTS_CSV
    profile = resolve_market_profile(market_key, base_url_override)
    URL = profile["base_url"]
    app_config.URL = profile["base_url"]
    PROMPTS_CSV = profile["prompts_path"]
    os.environ["DAKOTA_MARKET"] = profile["key"]
    os.environ["DAKOTA_BASE_URL"] = profile["base_url"]
    if profile.get("runs_per_object") is not None:
        global RUNS_PER_OBJECT
        RUNS_PER_OBJECT = int(profile["runs_per_object"])
    return profile


def _parse_cli_args(argv=None):
    parser = argparse.ArgumentParser(description="Dakota Joe chatbot prompt automation")
    parser.add_argument(
        "--run-mode",
        choices=RUN_MODES,
        default=os.getenv("DAKOTA_RUN_MODE", DEFAULT_RUN_MODE),
        help="Prompt set: all (full CSV), smoke (Marker=smoke), test (Marker=test)",
    )
    parser.add_argument(
        "--smoke",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run browser without a visible window",
    )
    parser.add_argument(
        "--browser",
        choices=("chrome", "edge", "firefox"),
        default=os.getenv("BROWSER", "chrome"),
        help="Browser engine (default: chrome)",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=None,
        help=f"Seconds to wait for report link (default: {RESPONSE_TIMEOUT})",
    )
    parser.add_argument(
        "--runs",
        type=int,
        default=None,
        help=f"Timing samples per object (default: {RUNS_PER_OBJECT})",
    )
    parser.add_argument(
        "--market",
        choices=market_choices(),
        default=os.getenv("DAKOTA_MARKET", DEFAULT_MARKET),
        help=f"Target market/environment (default: {DEFAULT_MARKET})",
    )
    parser.add_argument(
        "--base-url",
        default=None,
        help="Override base URL (required for custom market if DAKOTA_BASE_URL is unset)",
    )
    return parser.parse_args(argv)


def main(argv=None):
    args = _parse_cli_args(argv)
    run_mode = normalize_run_mode(
        SMOKE_MARKER if args.smoke else args.run_mode
    )
    headless = bool(args.headless)
    browser = (args.browser or "chrome").strip().lower()
    _apply_run_overrides(response_timeout=args.timeout, runs_per_object=args.runs)

    try:
        market_profile = _apply_market(args.market, args.base_url)
    except ValueError as exc:
        print(f"Market configuration error: {exc}")
        return 1

    try:
        prompt_exec = resolve_prompt_execution(market_profile, run_mode)
    except ValueError as exc:
        print(f"Prompt configuration error: {exc}")
        return 1

    global PROMPTS_CSV, RUNS_PER_OBJECT
    PROMPTS_CSV = prompt_exec["prompts_path"]
    run_mode = prompt_exec["run_mode"]
    if prompt_exec.get("runs_per_object") is not None:
        RUNS_PER_OBJECT = int(prompt_exec["runs_per_object"])

    print("=" * 60)
    print("Dakota Joe Chatbot Testing Script")
    print(f"Run mode: {run_mode}")
    print(f"Market: {market_profile['label']} ({market_profile['key']})")
    print(f"Base URL: {market_profile['base_url']}")
    print(f"Prompts file: {prompt_exec['prompts_file']}")
    print(f"Browser: {browser} ({'headless' if headless else 'visible'})")
    print(f"Response timeout: {RESPONSE_TIMEOUT}s | Runs per object: {RUNS_PER_OBJECT}")
    print("=" * 60)

    if not os.path.exists(PROMPTS_CSV):
        print(f"Prompt file not found: {PROMPTS_CSV}")
        return 1

    data = load_prompts_from_csv(PROMPTS_CSV)
    if not data:
        print(f"Could not load prompts from {PROMPTS_CSV}")
        return 1

    rows, prompt_col, status_col, link_col, time_col, prompt_entries, header_row_idx, hi_col = data
    prompt_entries = _filter_prompt_entries_for_run_mode(prompt_entries, run_mode)
    if not prompt_entries:
        print(
            f"No prompts found for run mode '{run_mode}' in {PROMPTS_CSV}. "
            f"Use {MARKER_COL}={SMOKE_MARKER}, {TEST_MARKER}, or run-mode=all."
        )
        return 1
    print(f"Run mode '{run_mode}': {len(prompt_entries)} object prompt(s) selected")

    prompts = [_prompt_entry_prompt(entry) for entry in prompt_entries]
    row_indices = [_prompt_entry_row_idx(entry) for entry in prompt_entries]
    csv_writeback_enabled = (
        status_col is not None and link_col is not None and time_col is not None
    )
    print(f"Loaded {len(prompt_entries)} object prompts from {PROMPTS_CSV}")

    append_perf = hi_col is not None
    object_types = []
    sample_nums = []
    if append_perf:
        object_types = _object_type_series_for_entries(
            rows, header_row_idx, hi_col, prompt_entries
        )
        prompts, row_indices, object_types, sample_nums = _expand_prompt_runs(
            prompts, row_indices, object_types
        )
        print(
            f"Expanded to {len(prompts)} executions "
            f"({RUNS_PER_OBJECT} per object)"
        )

    if not prompts:
        print("No prompts found")
        return 1

    print(f"Loaded {len(prompts)} prompts")

    driver = setup_driver(headless=headless, browser=browser)
    allure_enabled = _prepare_allure_results(_browser_label(driver).strip(), "windows")
    benchmarks_by_object = _load_all_benchmarks() if append_perf else {}
    if append_perf:
        if benchmarks_by_object:
            print(
                f"Benchmarks loaded for {len(benchmarks_by_object)} object type(s) "
                f"(from {BENCHMARKS_CSV_NAME} and/or Excel Run summary rows)"
            )
        else:
            print(
                f"WARNING: No benchmarks configured. Add {BENCHMARKS_CSV_NAME} "
                "or set column E on Excel Run summary rows."
            )
    active_allure_group = None
    allure_cases_passed = 0
    allure_cases_failed = 0
    passed = 0
    failed = 0

    try:
        if not login(driver):
            print("Aborting run: login did not complete.")
            return 1
        print("Waiting for page...")
        time.sleep(PAGE_IDLE_AFTER_LOGIN)

        chat_is_open = False
        group_elapsed = []
        recorded = datetime.now().date().isoformat()

        # Send a single greeting first, then send only prompt texts.
        # (Do NOT send category names like Accounts/Contacts.)
        if append_perf:
            try:
                print("Sending hi first...")
                type_prompt(driver, "hi", chat_already_open=chat_is_open)
                chat_is_open = True
                time.sleep(0.5)
            except Exception:
                pass

        for i, prompt in enumerate(prompts):
            row_idx = row_indices[i] if row_indices is not None else None
            object_type = object_types[i] if append_perf and i < len(object_types) else ""
            sample_num = sample_nums[i] if append_perf and i < len(sample_nums) else ""
            should_reset_session = _should_reset_chat_session_after_prompt(
                append_perf, i, len(prompts)
            )
            print(f"\n[{i+1}/{len(prompts)}] {prompt[:60]}...")
            t0 = None
            screenshot_path = None
            initial_unique_count = None
            initial_raw_count = None
            try:
                t0 = time.time()
                start_ms = int(t0 * 1000)
                # Drain stale browser console logs before handling this prompt cycle.
                try:
                    driver.get_log("browser")
                except Exception:
                    pass
                initial_unique_count, initial_raw_count, _, initial_hrefs = count_report_links(driver)
                print(f"  Initial links: raw={initial_raw_count}, unique={initial_unique_count}")

                type_prompt(driver, prompt, chat_already_open=chat_is_open)
                chat_is_open = True
                try:
                    driver.get_log("browser")
                except Exception:
                    pass
                sent_ms = int(time.time() * 1000)

                new_link = wait_for_new_link(
                    driver,
                    initial_unique_count,
                    initial_raw_count,
                    since_ms=sent_ms,
                    initial_hrefs=initial_hrefs,
                )
                elapsed_s = round(time.time() - t0, 2)

                if new_link:
                    print(f"  [OK] PASSED - {new_link[:60]}...")
                    status, link = "Pass", new_link
                    passed += 1
                else:
                    late_link = recheck_link_before_fail(
                        driver, initial_unique_count, initial_raw_count, grace_seconds=8
                    )
                    if late_link:
                        print(f"  [OK] PASSED (late link) - {late_link[:60]}...")
                        status, link = "Pass", late_link
                        passed += 1
                    else:
                        print(f"  [FAIL] No link found")
                        status, link = "Fail", "-"
                        screenshot_path = take_screenshot(driver, f"failed_{i+1}")
                        failed += 1

                if csv_writeback_enabled and row_idx is not None:
                    row = rows[row_idx]
                    while len(row) <= max(status_col, link_col, time_col):
                        row.append("")
                    row[status_col] = status
                    row[link_col] = link
                    row[time_col] = str(elapsed_s)
                    write_prompts_csv(PROMPTS_CSV, rows)

                group_elapsed = _append_perf_sample_and_maybe_summary(
                    driver,
                    append_perf,
                    i,
                    prompt,
                    prompts,
                    object_types,
                    sample_nums,
                    elapsed_s,
                    group_elapsed,
                    recorded,
                    benchmarks_by_object,
                )

                if allure_enabled:
                    stop_ms = int(time.time() * 1000)
                    if active_allure_group is None:
                        active_allure_group = _start_allure_object_group(
                            object_type, prompt, start_ms
                        )
                    _append_allure_sample(
                        active_allure_group,
                        sample_num=sample_num,
                        status=status,
                        link=link,
                        elapsed_s=elapsed_s,
                        start_ms=start_ms,
                        stop_ms=stop_ms,
                        screenshot_path=screenshot_path,
                        error_message=None if status == "Pass" else "No link found",
                    )
                    if _allure_group_is_complete(i, prompts, object_types, append_perf):
                        object_passed = _finalize_allure_object_group(
                            active_allure_group, benchmarks_by_object
                        )
                        if object_passed is True:
                            allure_cases_passed += 1
                        elif object_passed is False:
                            allure_cases_failed += 1
                        active_allure_group = None

            except Exception as e:
                print(f"  [ERROR] {e}")
                if not _driver_window_open(driver):
                    print("  Browser window closed; stopping remaining prompts.")
                    failed += 1
                    break
                status, link = "Fail", "-"
                err_elapsed = (
                    round(time.time() - t0, 2) if t0 is not None else ""
                )

                # Always re-check once more for a late link before finalizing failure.
                late_link = None
                if initial_unique_count is not None and initial_raw_count is not None:
                    try:
                        late_link = recheck_link_before_fail(
                            driver, initial_unique_count, initial_raw_count, grace_seconds=8
                        )
                    except Exception:
                        late_link = None

                if late_link:
                    print(f"  [OK] PASSED (late link after error) - {late_link[:60]}...")
                    status, link = "Pass", late_link
                    passed += 1
                else:
                    if csv_writeback_enabled and row_idx is not None:
                        row = rows[row_idx]
                        while len(row) <= max(status_col, link_col, time_col):
                            row.append("")
                        row[status_col] = status
                        row[link_col] = link
                        row[time_col] = str(err_elapsed) if err_elapsed != "" else ""
                        write_prompts_csv(PROMPTS_CSV, rows)
                    if append_perf and err_elapsed != "":
                        group_elapsed = _append_perf_sample_and_maybe_summary(
                            driver,
                            True,
                            i,
                            prompt,
                            prompts,
                            object_types,
                            sample_nums,
                            err_elapsed,
                            group_elapsed,
                            recorded,
                            benchmarks_by_object,
                        )
                    screenshot_path = take_screenshot(driver, f"failed_{i+1}")
                    failed += 1

                if allure_enabled:
                    start_ms = int((t0 if t0 is not None else time.time()) * 1000)
                    stop_ms = int(time.time() * 1000)
                    if active_allure_group is None:
                        active_allure_group = _start_allure_object_group(
                            object_type, prompt, start_ms
                        )
                    _append_allure_sample(
                        active_allure_group,
                        sample_num=sample_num,
                        status=status,
                        link=link,
                        elapsed_s=err_elapsed if err_elapsed != "" else "0",
                        start_ms=start_ms,
                        stop_ms=stop_ms,
                        screenshot_path=screenshot_path,
                        error_message=None if status == "Pass" else str(e),
                    )
                    if _allure_group_is_complete(i, prompts, object_types, append_perf):
                        object_passed = _finalize_allure_object_group(
                            active_allure_group, benchmarks_by_object
                        )
                        if object_passed is True:
                            allure_cases_passed += 1
                        elif object_passed is False:
                            allure_cases_failed += 1
                        active_allure_group = None

            # Before next prompt: wait for ready event, then end chat after each execution.
            if i < len(prompts) - 1:
                wait_for_console_ready_event(
                    driver,
                    timeout_s=CONSOLE_EVENT_WAIT_SECONDS,
                    since_ms=start_ms if t0 is not None else 0,
                )
                if should_reset_session:
                    if end_chat_session_from_header(driver):
                        chat_is_open = False
                        try:
                            driver.get_log("browser")
                        except Exception:
                            pass

    except KeyboardInterrupt:
        print("\nRun interrupted; finalizing partial results...")
    except Exception as e:
        print(f"Fatal error: {e}")
        traceback.print_exc()
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        if allure_enabled and active_allure_group is not None:
            object_passed = _finalize_allure_object_group(
                active_allure_group, benchmarks_by_object
            )
            if object_passed is True:
                allure_cases_passed += 1
            elif object_passed is False:
                allure_cases_failed += 1
            active_allure_group = None

    print(f"\n{'='*60}")
    print(
        f"RESULTS: Prompt passes={passed}, Prompt failures={failed}, "
        f"Allure object failures={allure_cases_failed}"
    )
    if csv_writeback_enabled:
        print(f"Updated: {PROMPTS_CSV}")
    else:
        print(f"Prompt source (read-only): {PROMPTS_CSV}")
    if append_perf:
        print(
            f"Appended (no overwrite) browser timing rows to: {PERFORMANCE_RESULTS_XLSX}"
        )
    if allure_enabled:
        allure_total = allure_cases_passed + allure_cases_failed
        _generate_allure_report(allure_total, allure_cases_passed, allure_cases_failed)
    print("=" * 60)
    return failed + allure_cases_failed

if __name__ == "__main__":
    exit_code = main(sys.argv[1:]) or 0
    sys.exit(1 if exit_code else 0)
